import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import fitz  # PyMuPDF


def get_pdf_page_count(path: str) -> int:
    with fitz.open(path) as doc:
        return doc.page_count


def save_extraction_result_xlsx(
    data: dict, total_time: float, file_path: str, output_dir="results"
):
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "extraction_results.xlsx")

    # --- Подготовка строки данных ---
    file_name = os.path.basename(file_path)
    row = {
        "file_name": file_name,
        "file_path": file_path,
        **data,
        "page_count": get_pdf_page_count(file_path),
        "processing_time_sec": round(total_time, 2),
    }

    df = pd.DataFrame([row])

    # --- Если файл уже существует — добавляем строку ---
    if os.path.exists(output_path):
        existing_df = pd.read_excel(output_path)
        df = pd.concat([existing_df, df], ignore_index=True)

    # --- Сохраняем в Excel ---
    df.to_excel(output_path, index=False, engine="openpyxl")

    # --- Применяем форматирование ---
    wb = load_workbook(output_path)
    ws = wb.active

    # Заголовки жирным и по центру
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Применяем формат к ячейкам
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align_center
            cell.border = thin_border

    # Заголовки — жирные
    for cell in ws[1]:
        cell.font = header_font

    # Автоматическая ширина столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(output_path)
    wb.close()

    print(f"Excel файл успешно обновлён: {output_path}")
    return output_path
